#!/usr/bin/env python
import os
import sys
import time
import re
import argparse
import openpyxl
from openpyxl.utils import *
from openpyxl.styles import Font
#HOME = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))

def cli_parser():
    parser = argparse.ArgumentParser(description = "Log Format to CSV")
    parser.add_argument('-b', '--board', required=True, help ="board type")
    parser.add_argument('-v', '--version', required=True, help = "board version")
    parser.add_argument('-l', '--log', required=True, help = "log file")
    args = parser.parse_args()
    return args

def log_show_temp(log):
    diagp = r'Diag\D+\S+\s\d*'
    start = r"ID\s+Sensor\s+Temperature*$"
    pattern = r"\d+\s*(?P<sensor>[a-zA-Z_0-9]*)\s+(?P<temp>[\d+.]+)"
    temp_dict = {}

    with open(log, 'r') as f:
        data = f.read()
    strs = data.split('\n')
    for i in range(len(strs)):
        begin = re.search(start,strs[i])
        if begin is not None:
            for st in range(i,len(strs)):
                end = re.search(diagp,strs[st])
                if end is None:
                    temp = re.search(pattern, strs[st])
                    if temp is not None:
                        temp_dict[temp.group(1)]=temp.group(2)
                    else:
                        continue
                else:
                    continue
    return temp_dict

def log_xcvr_temp(log):
    begin_pat = re.compile(r"Diag.*transceiver\sshow_temp")
    match_pat1 = re.compile(r"(Transceiver\W+(\d+).*(Q.{6,20})[:])|(.*inst\s(\d+))")
    match_pat2 = re.compile(r"Transceiver\D(\d+)\W+sensor\D+(\d)\D+(\d+[.]\d)")
    matched_qsfp = re.compile(r"Transceiver\W(\d+)\W+temperature\W+(\d+[.]\d)")
    error_pat = re.compile(r"ERROR\D+(\d+).*(not.*)")
    match_item_dict = {}
    match2_dict = {}
    dict_qsfp = {}
    dict_err = {}
    matching_temp_dict = {}
        
    with open(log, 'r') as f:
        data = f.read()
    strs = data.split('\n')
    for i in range(len(strs)):
        begin = begin_pat.search(strs[i])
        if begin is not None:
            for n in range(i, len(strs)):  
                match1 = match_pat1.search(strs[n])
                match2 = match_pat2.search(strs[n])
                qsfp = matched_qsfp.search(strs[n])
                errs = error_pat.search(strs[n])
                if match1 is not None:
                    inst = match1.group(2)
                    if inst is None:
                        inst = match1.group(5)
                    match_item_dict[inst] = match1.group(3)  ##match module type
                    matching_temp_dict[inst] = {}
                if match2 is not None:
                    inst = match2.group(1)
                    sensor_no = match2.group(2)
                    temp = match2.group(3)
                    matching_temp_dict[inst][sensor_no] = temp
                if qsfp is not None:
                    dict_qsfp[qsfp.group(1)] = qsfp.group(2)
                if errs is not None:
                    dict_err[errs.group(1)] = errs.group(2)
    
    for key in match_item_dict.keys():
        if key in dict_qsfp.keys():
            matching_temp_dict[key]= dict_qsfp[key]
        if key in dict_err.keys():
            matching_temp_dict[key] = dict_err[key]
                
    return match_item_dict, matching_temp_dict
      
def workbook_creation(board_name, board_version,dict1,dict_xcvr, dict_value):
    wb = openpyxl.Workbook()
    # workbook settings
    date = str(time.strftime("%Y_%m_%d_%H_%M"))
    wb_title = board_name + '_'+ board_version + "_Thermal Test" + '_' + date + '.xlsx'
    ws1 = wb.active
    width = 25.0
    height = 40.0
    start_row = 3
    for i in range(1,4):
        ws1.row_dimensions[i].height = height
    for i in range(1,10):
        ws1.column_dimensions[get_column_letter(i)].width = width
        
    ws1.title = "Temperature_Sweep Diag"     
    ws1['A1'] = "Thermal couple/ Sensor"
    ws1['B1'] = "Components"
    ws1['C1'] = "Power(W)"
    ws1['D1'] = "Temp. Rating"
    ws1['E1'] = "Case/Junction"
    ws1['F1'] = "Sensor/TC Location"
    ws1['G1'] = "Measured Temp"
    ws1['H1'] = "Margin"
    font = Font(size = 14, bold = True)
    for cell in ws1[1]:
        cell.font = font

#data write
    i = 0
    for key,value in dict1.items():
        ws1.cell(row = start_row+i, column = 2, value = key)
        ws1.cell(row = start_row+i, column = 7, value = value)
        i = i+1       
    n = i   # begin no of xcvr
    for key,value in dict_xcvr.items():
        ws1.cell(row = start_row+i, column = 2, value = value)
        i = i+1
    end_line = i #end line of whole items
    num =0
    for key,value in dict_value.items():
        if type(value) is not dict:
            ws1.cell(row = start_row+n+num, column = 7, value = value)
        else:
            ws1.cell(row = start_row+n+num, column = 7, value = str(value))
        num = num+1
       
    for m in range(end_line):
        ws1.cell(row = start_row+m, column = 1, value = m)
    for m in range(end_line):
        ws1.cell(row = start_row+m, column = 5, value = "Junciton" )
    for m in range(n, end_line):
        ws1.cell(row = start_row+m, column = 5, value ="Case")
    for m in range(end_line):
        ws1.cell(row = start_row+m, column = 6,value = "Internal Sensors")

    wb.save(wb_title)
    return wb
    

if __name__ == "__main__":
    args = cli_parser()
    board_type = args.board
    board_version = args.version
    log = args.log
    dict_temp = log_show_temp(log)
    dict_mods,dict_values = log_xcvr_temp(log)
    workbook_creation(board_type,board_version,dict_temp,dict_mods,dict_values)                 
